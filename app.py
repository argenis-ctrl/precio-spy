"""
Dashboard de Monitoreo de Precios - Depilación Láser
Lasertam vs. Competencia | Chile
"""

import io
import json
import sqlite3
import sys
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
import plotly.express as px
import streamlit as st

sys.path.insert(0, str(Path(__file__).parent))
from db.models import DB_PATH, init_db

# ── Cupones conocidos ──────────────────────────────────────────────────────
_COUPONS_PATH = Path(__file__).parent / "coupons.json"

def load_coupons() -> dict:
    """Retorna {competitor: best_discount_pct} para aplicar al precio."""
    if not _COUPONS_PATH.exists():
        return {}
    try:
        data = json.loads(_COUPONS_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {}
    result = {}
    for competitor, coupons in data.items():
        best = 0
        for c in coupons:
            if c.get("type") == "pct":
                best = max(best, c["value"])
        if best > 0:
            result[competitor] = best
    return result

COUPONS = load_coupons()  # {"Cela": 20, ...}

# ── Configuración de página ────────────────────────────────────────────────
st.set_page_config(
    page_title="PrecioSpy · Lasertam",
    page_icon="./static/favicon.svg",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── CSS global ─────────────────────────────────────────────────────────────
st.markdown("""
<style>
/* Ocultar elementos de Streamlit */
#MainMenu, footer, [data-testid="stDecoration"] { visibility: hidden; }

/* ── Sidebar oscuro ── */
[data-testid="stSidebar"] {
    background: #0f172a !important;
    border-right: 1px solid #1e293b;
}
[data-testid="stSidebar"] .stMarkdown,
[data-testid="stSidebar"] .stCaption,
[data-testid="stSidebar"] label,
[data-testid="stSidebar"] p,
[data-testid="stSidebar"] span,
[data-testid="stSidebar"] small { color: #94a3b8 !important; }
[data-testid="stSidebar"] h1,
[data-testid="stSidebar"] h2,
[data-testid="stSidebar"] h3 { color: #f1f5f9 !important; }
[data-testid="stSidebar"] hr { border-color: #1e293b !important; }
[data-testid="stSidebar"] .stRadio label { color: #cbd5e1 !important; }
[data-testid="stSidebar"] [data-testid="stToggle"] span { color: #cbd5e1 !important; }

/* ── Tabs limpios ── */
.stTabs [data-baseweb="tab-list"] {
    gap: 0;
    border-bottom: 2px solid #e2e8f0;
    background: transparent;
}
.stTabs [data-baseweb="tab"] {
    padding: 10px 22px;
    font-size: 13px;
    font-weight: 500;
    color: #64748b;
    background: transparent;
    border-bottom: 2px solid transparent;
    margin-bottom: -2px;
    border-radius: 0;
}
.stTabs [data-baseweb="tab"]:hover { color: #0ea5e9; background: transparent; }
.stTabs [aria-selected="true"] {
    color: #0ea5e9 !important;
    border-bottom: 2px solid #0ea5e9 !important;
    background: transparent !important;
    font-weight: 600;
}

/* ── Tarjetas KPI ── */
[data-testid="metric-container"] {
    background: white;
    border: 1px solid #e2e8f0;
    border-radius: 10px;
    padding: 18px 20px;
    box-shadow: 0 1px 4px rgba(0,0,0,0.04);
}
[data-testid="stMetricValue"] { font-size: 26px !important; font-weight: 700; color: #0f172a !important; }
[data-testid="stMetricLabel"] { font-size: 12px !important; font-weight: 500; color: #64748b !important; text-transform: uppercase; letter-spacing: .04em; }

/* ── Header de página ── */
.page-header {
    padding: 10px 0 18px 0;
    border-bottom: 1px solid #e2e8f0;
    margin-bottom: 24px;
}
.page-header h1 {
    font-size: 22px;
    font-weight: 700;
    color: #0f172a;
    margin: 0;
    letter-spacing: -0.3px;
}
.page-header p {
    font-size: 13px;
    color: #64748b;
    margin: 4px 0 0 0;
}

/* ── Logo sidebar ── */
.sidebar-brand {
    padding: 8px 0 16px 0;
    border-bottom: 1px solid #1e293b;
    margin-bottom: 8px;
}
.sidebar-brand-name {
    font-size: 18px;
    font-weight: 800;
    color: #f1f5f9 !important;
    letter-spacing: -0.5px;
}
.sidebar-brand-tag {
    font-size: 10px;
    text-transform: uppercase;
    letter-spacing: .1em;
    color: #0ea5e9 !important;
    font-weight: 600;
}

/* ── Botón actualizar ── */
[data-testid="stSidebar"] .stButton button,
[data-testid="stSidebar"] .stButton > button,
[data-testid="stSidebar"] button[kind="primary"] {
    background: #0ea5e9 !important;
    color: #ffffff !important;
    border: none !important;
    border-radius: 7px !important;
    font-weight: 600 !important;
    font-size: 13px !important;
}
[data-testid="stSidebar"] .stButton button p,
[data-testid="stSidebar"] .stButton button span {
    color: #ffffff !important;
}
[data-testid="stSidebar"] .stButton button:hover {
    background: #0284c7 !important;
}

/* ── Expander ── */
[data-testid="stExpander"] {
    border: 1px solid #e2e8f0 !important;
    border-radius: 10px !important;
    background: white;
    box-shadow: 0 1px 3px rgba(0,0,0,0.04);
}
</style>
""", unsafe_allow_html=True)

COMPETITOR_COLORS = {
    "Belenus":      "#ff7f0e",
    "Cela":         "#2ca02c",
    "Bellmeclinic": "#d62728",
    "Lasertam":     "#1f77b4",
}

COMPANIES_ORDER = ["Lasertam", "Belenus", "Cela", "Bellmeclinic"]
CO_SHORT  = {"Lasertam": "Lasertam", "Belenus": "Belenus", "Cela": "Cela", "Bellmeclinic": "Bellme."}
CO_HEADER = {"Lasertam": "#dbeafe", "Belenus": "#fff3e0", "Cela": "#e8f5e9", "Bellmeclinic": "#fce4ec"}
CO_TEXT   = {"Lasertam": "#1e40af", "Belenus": "#e65100", "Cela": "#1b5e20", "Bellmeclinic": "#880e4f"}

# ── Helpers de DB ──────────────────────────────────────────────────────────

@st.cache_resource
def get_conn():
    init_db()
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    return conn


def run_query(sql: str, params=()) -> pd.DataFrame:
    conn = get_conn()
    return pd.read_sql_query(sql, conn, params=params)


# ── Carga de datos ─────────────────────────────────────────────────────────

@st.cache_data(ttl=300)
def load_latest_prices(gender_filter: str, active_coupons: tuple = ()) -> pd.DataFrame:
    """active_coupons: tuple de (competitor, pct) para los cupones activados."""
    gender_clause = "AND pr.gender = ?" if gender_filter != "Todos" else ""
    q = f"""
    SELECT
        c.name                  AS competitor,
        c.is_self,
        pr.zone_name,
        pr.gender,
        pr.sessions,
        MIN(pr.price)           AS price,
        MIN(pr.original_price)  AS original_price,
        MAX(pr.discount_pct)    AS discount_pct,
        MAX(pr.scraped_at)      AS scraped_at
    FROM price_records pr
    JOIN competitors c ON c.id = pr.competitor_id
    WHERE pr.scraped_at = (
        -- Último precio conocido por zona (independiente del run)
        SELECT MAX(pr2.scraped_at)
        FROM price_records pr2
        WHERE pr2.competitor_id = pr.competitor_id
          AND pr2.zone_name     = pr.zone_name
          AND pr2.gender        = pr.gender
          AND pr2.sessions      IS pr.sessions
    )
    {gender_clause}
    GROUP BY c.name, pr.zone_name, pr.gender, pr.sessions
    """
    params = []
    if gender_filter != "Todos":
        params.append(gender_filter[0])
    df = run_query(q, params)

    # Aplicar solo los cupones que el usuario tiene activados
    active_map = dict(active_coupons)
    if not df.empty and active_map:
        for competitor, pct in active_map.items():
            mask = df["competitor"] == competitor
            if not mask.any():
                continue
            factor = 1 - pct / 100
            # Si no había precio original, el precio listado pasa a ser el "normal"
            no_orig = mask & df["original_price"].isna()
            df.loc[no_orig, "original_price"] = df.loc[no_orig, "price"]
            # Precio con cupón
            df.loc[mask, "price"] = (df.loc[mask, "price"] * factor).round(0).astype(int)
            # Recalcular descuento total (precio final vs precio normal)
            df.loc[mask, "discount_pct"] = (
                (1 - df.loc[mask, "price"] / df.loc[mask, "original_price"]) * 100
            ).round(1)
            df.loc[mask, "has_coupon"] = True

    if "has_coupon" not in df.columns:
        df["has_coupon"] = False

    return df


@st.cache_data(ttl=300)
def load_price_history(zone: str, gender: str, sessions) -> pd.DataFrame:
    q = """
    SELECT c.name AS competitor, pr.price, pr.sessions, DATE(pr.scraped_at) AS fecha
    FROM price_records pr
    JOIN competitors c ON c.id = pr.competitor_id
    WHERE pr.zone_name = ? AND pr.gender = ?
    """
    params = [zone, gender]
    if sessions:
        q += " AND pr.sessions = ?"
        params.append(sessions)
    q += " ORDER BY pr.scraped_at"
    return run_query(q, params)


@st.cache_data(ttl=300)
def load_scrape_dates() -> pd.DataFrame:
    return run_query("""
        SELECT c.name, MAX(pr.scraped_at) AS ultimo_scrape
        FROM price_records pr
        JOIN competitors c ON c.id = pr.competitor_id
        GROUP BY c.name
    """)


# ── Utilidades de formato ──────────────────────────────────────────────────

def fmt_clp(val) -> str:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return "—"
    return f"${int(val):,}".replace(",", ".")


# ── Tabla HTML comparativa ─────────────────────────────────────────────────

TABLE_CSS = """
<style>
.cmp-wrap { overflow-x: auto; margin-top: 8px; }
.cmp-table {
    border-collapse: collapse;
    width: 100%;
    font-size: 12.5px;
    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
}
.cmp-table th, .cmp-table td {
    border: 1px solid #e5e7eb;
    padding: 6px 8px;
    white-space: nowrap;
    text-align: center;
    vertical-align: middle;
}
.cmp-table .zone-col {
    text-align: left;
    font-weight: 600;
    min-width: 150px;
    background: #f9fafb;
    position: sticky;
    left: 0;
    z-index: 1;
    border-right: 2px solid #d1d5db;
}
.cmp-table .ses-hdr {
    background: #f3f4f6;
    font-weight: 700;
    font-size: 13px;
    border-bottom: 2px solid #d1d5db;
    padding: 8px 4px;
}
.co-hdr {
    font-weight: 700;
    font-size: 11px;
    padding: 4px 6px !important;
    border-bottom: 2px solid;
}
.cmp-table .price-cell { min-width: 95px; padding: 5px 8px; }
.cmp-table .no-data { color: #d1d5db; font-size: 11px; }
.orig { text-decoration: line-through; color: #9ca3af; font-size: 10.5px; display: block; line-height: 1.2; }
.offer { font-weight: 700; font-size: 13.5px; color: #111827; display: block; line-height: 1.4; }
.disc-badge {
    display: inline-block;
    background: #dcfce7;
    color: #15803d;
    border-radius: 4px;
    padding: 0px 5px;
    font-size: 10px;
    font-weight: 700;
    margin-top: 1px;
}
.coupon-badge {
    display: inline-block;
    background: #fef9c3;
    color: #854d0e;
    border-radius: 4px;
    padding: 0px 5px;
    font-size: 9px;
    font-weight: 700;
    margin-left: 2px;
    border: 1px solid #fde68a;
}
.cheapest { background: #f0fdf4 !important; }
.most-exp { background: #fff1f2 !important; }
.zebra { background: #fafafa; }
.zone-head-row th { border-top: 3px solid #e5e7eb; }
</style>
"""


def build_comparison_table(df: pd.DataFrame, sessions_list: list, search: str = "") -> str:
    """Genera tabla HTML comparativa zona × empresa × sesiones."""

    companies = [c for c in COMPANIES_ORDER if c in df["competitor"].values]
    if not companies:
        return "<p>Sin datos.</p>"

    # Lookup rápido: (zone, company, sessions) → dict
    lookup: dict = {}
    for _, r in df.iterrows():
        ses = int(r["sessions"]) if pd.notna(r["sessions"]) else None
        key = (r["zone_name"], r["competitor"], ses)
        cur = lookup.get(key)
        if cur is None or r["price"] < cur["price"]:
            lookup[key] = {
                "price":      r["price"],
                "original":   r["original_price"] if pd.notna(r.get("original_price")) else None,
                "discount":   r["discount_pct"]   if pd.notna(r.get("discount_pct"))   else None,
                "has_coupon": bool(r.get("has_coupon", False)),
            }

    # Filtrar y ordenar zonas
    zones = sorted(df["zone_name"].unique())
    if search:
        zones = [z for z in zones if search.lower() in z.lower()]

    def make_cell(zone, company, ses):
        d = lookup.get((zone, company, ses))
        if not d:
            return '<td class="no-data price-cell">—</td>'

        price      = d["price"]
        orig       = d["original"]
        disc       = d["discount"]
        has_coupon = d.get("has_coupon", False)

        # Hallar si es el más barato o el más caro en esta zona×sesión
        prices_here = [lookup[(zone, co, ses)]["price"]
                       for co in companies if (zone, co, ses) in lookup]
        is_cheapest  = prices_here and price == min(prices_here)
        is_most_exp  = prices_here and price == max(prices_here) and len(prices_here) > 1

        cell_class = "price-cell cheapest" if is_cheapest else (
                     "price-cell most-exp" if is_most_exp else "price-cell")

        inner = ""
        if orig and orig > price:
            inner += f'<span class="orig">{fmt_clp(orig)}</span>'
        inner += f'<span class="offer">{fmt_clp(price)}</span>'
        if disc and disc > 0:
            inner += f'<span class="disc-badge">-{int(disc)}%</span>'
        if has_coupon and company in COUPONS:
            inner += f'<span class="coupon-badge">cupón</span>'

        return f'<td class="{cell_class}">{inner}</td>'

    # ── Construir HTML ──────────────────────────────────────────────────────
    html = TABLE_CSS + '<div class="cmp-wrap"><table class="cmp-table">'

    # Fila 1: sesiones (spanning)
    html += '<thead><tr class="zone-head-row">'
    html += '<th class="zone-col" rowspan="2">Zona</th>'
    for ses in sessions_list:
        lbl = f"{ses} ses." if ses else "Paquete"
        html += f'<th class="ses-hdr" colspan="{len(companies)}">{lbl}</th>'
    html += '</tr>'

    # Fila 2: empresas
    html += '<tr>'
    for _ in sessions_list:
        for co in companies:
            bg   = CO_HEADER.get(co, "#f3f4f6")
            clr  = CO_TEXT.get(co,   "#111827")
            brd  = CO_TEXT.get(co,   "#9ca3af")
            name = CO_SHORT.get(co, co)
            html += (f'<th class="co-hdr" '
                     f'style="background:{bg};color:{clr};border-bottom-color:{brd}">'
                     f'{name}</th>')
    html += '</tr></thead><tbody>'

    # Filas de datos
    for i, zone in enumerate(zones):
        row_class = "zebra" if i % 2 == 0 else ""
        html += f'<tr class="{row_class}">'
        html += f'<td class="zone-col">{zone}</td>'
        for ses in sessions_list:
            for co in companies:
                html += make_cell(zone, co, ses)
        html += '</tr>'

    html += '</tbody></table></div>'
    return html


# ── Excel export ───────────────────────────────────────────────────────────

def generate_excel_report(df: pd.DataFrame) -> bytes:
    """Genera reporte Excel: hoja Comparativa + hoja % vs Lasertam."""
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    output = io.BytesIO()
    companies = [c for c in COMPANIES_ORDER if c in df["competitor"].values]

    # Pivot plano: una fila por (zona, sesiones), columnas = empresas
    pivot = (
        df.pivot_table(
            index=["zone_name", "sessions"],
            columns="competitor",
            values="price",
            aggfunc="min",
        )
        .reset_index()
    )
    pivot.columns.name = None
    pivot["sessions"] = pivot["sessions"].apply(
        lambda s: f"{int(s)} ses." if pd.notna(s) else "Paquete"
    )
    pivot = pivot.rename(columns={"zone_name": "Zona", "sessions": "Sesiones"})
    col_order = ["Zona", "Sesiones"] + [c for c in companies if c in pivot.columns]
    pivot = pivot[col_order].sort_values(["Zona", "Sesiones"])

    # Estilos reutilizables
    dark  = PatternFill("solid", fgColor="0F172A")
    green = PatternFill("solid", fgColor="DCFCE7")
    red   = PatternFill("solid", fgColor="FEE2E2")
    hdr_font  = Font(color="FFFFFF", bold=True, size=10)
    g_font    = Font(color="166534", bold=True)
    r_font    = Font(color="991B1B", bold=True)
    center    = Alignment(horizontal="center")

    with pd.ExcelWriter(output, engine="openpyxl") as writer:

        # ── Hoja 1: Comparativa ─────────────────────────────────────────────
        pivot.to_excel(writer, sheet_name="Comparativa de Precios", index=False)
        ws = writer.sheets["Comparativa de Precios"]

        for cell in ws[1]:
            cell.fill, cell.font, cell.alignment = dark, hdr_font, center

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 12
        for i in range(3, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(i)].width = 16

        # Índices de columna para cada empresa
        co_col = {co: col_order.index(co) + 1 for co in companies if co in col_order}

        for row_idx in range(2, ws.max_row + 1):
            prices = {}
            for co, ci in co_col.items():
                v = ws.cell(row_idx, ci).value
                if v is not None and isinstance(v, (int, float)) and not pd.isna(v):
                    prices[co] = v
            min_p = min(prices.values()) if prices else None
            max_p = max(prices.values()) if prices else None

            for co, ci in co_col.items():
                cell = ws.cell(row_idx, ci)
                if cell.value is None:
                    continue
                cell.number_format = '"$"#,##0'
                cell.alignment = center
                if len(prices) > 1:
                    if cell.value == min_p:
                        cell.fill = green
                    elif cell.value == max_p:
                        cell.fill = red

        ws.freeze_panes = "C2"
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

        # ── Hoja 2: % vs Lasertam ────────────────────────────────────────────
        if "Lasertam" in companies:
            competitors = [c for c in companies if c != "Lasertam"]
            pct_rows = []
            for _, row in pivot.iterrows():
                lt = row.get("Lasertam")
                r = {"Zona": row["Zona"], "Sesiones": row["Sesiones"]}
                for co in competitors:
                    co_p = row.get(co)
                    if pd.notna(lt) and pd.notna(co_p) and lt > 0:
                        r[f"vs {co} (%)"] = round((co_p - lt) / lt * 100, 1)
                    else:
                        r[f"vs {co} (%)"] = None
                pct_rows.append(r)

            pd.DataFrame(pct_rows).to_excel(writer, sheet_name="% vs Lasertam", index=False)
            ws2 = writer.sheets["% vs Lasertam"]

            for cell in ws2[1]:
                cell.fill, cell.font, cell.alignment = dark, hdr_font, center

            ws2.column_dimensions["A"].width = 25
            ws2.column_dimensions["B"].width = 12
            for i in range(3, ws2.max_column + 1):
                ws2.column_dimensions[get_column_letter(i)].width = 18

            for row in ws2.iter_rows(min_row=2):
                for cell in row:
                    if cell.column <= 2 or cell.value is None:
                        continue
                    if not isinstance(cell.value, (int, float)) or pd.isna(cell.value):
                        continue
                    cell.alignment = center
                    cell.number_format = '+0.0;-0.0;0'
                    if cell.value > 3:
                        cell.fill, cell.font = green, g_font
                    elif cell.value < -3:
                        cell.fill, cell.font = red, r_font

            ws2.freeze_panes = "C2"

    output.seek(0)
    return output.getvalue()


# ── Sidebar ────────────────────────────────────────────────────────────────

with st.sidebar:
    st.markdown("""
    <div class="sidebar-brand">
        <div class="sidebar-brand-name">PrecioSpy</div>
        <div class="sidebar-brand-tag">Inteligencia Competitiva</div>
    </div>
    """, unsafe_allow_html=True)

    gender_filter = st.radio("Género", ["Femenino", "Masculino", "Todos"], index=0)
    st.divider()

    # Cupones activos
    if COUPONS:
        st.caption("**Cupones de competencia:**")
        apply_coupons = {}
        for competitor, pct in COUPONS.items():
            apply_coupons[competitor] = st.toggle(
                f"{competitor} -{pct}%",
                value=True,
                help=f"Aplicar cupón conocido de {competitor} ({pct}% de descuento adicional)",
            )
        st.divider()
    else:
        apply_coupons = {}

    if st.button("Actualizar datos", type="primary", use_container_width=True):
        with st.spinner("Scrapeando sitios... (puede tardar ~10 min)"):
            try:
                import subprocess
                result = subprocess.run(
                    [sys.executable, "-m", "scraper.run_all"],
                    capture_output=True, text=True, timeout=700,
                )
                st.cache_data.clear()
                if result.returncode == 0:
                    st.success("Datos actualizados ✓")
                else:
                    st.warning(f"Completado con advertencias:\n{result.stderr[-500:]}")
            except Exception as e:
                st.error(f"Error: {e}")

    st.divider()
    dates_df = load_scrape_dates()
    if not dates_df.empty:
        st.caption("**Último scrape:**")
        for _, row in dates_df.iterrows():
            dt = row["ultimo_scrape"]
            if dt:
                try:
                    dt_parsed = datetime.fromisoformat(str(dt).replace("Z", "+00:00"))
                    dt_str = dt_parsed.strftime("%d/%m/%Y %H:%M")
                except Exception:
                    dt_str = str(dt)[:16]
                st.caption(f"• {row['name']}: {dt_str}")


# ── Main ───────────────────────────────────────────────────────────────────

st.markdown("""
<div class="page-header">
    <h1>Monitor de Precios — Depilación Láser</h1>
    <p>Lasertam &nbsp;·&nbsp; Belenus &nbsp;·&nbsp; Cela &nbsp;·&nbsp; Bellmeclinic &nbsp;·&nbsp; Chile</p>
</div>
""", unsafe_allow_html=True)

# Construir tuple de cupones activos (hasheable para el cache)
active_coupons_tuple = tuple(
    (c, pct) for c, pct in COUPONS.items()
    if apply_coupons.get(c, False)
)

df_all = load_latest_prices(gender_filter, active_coupons_tuple)

if df_all.empty:
    st.info("No hay datos aún. Haz clic en **Actualizar datos ahora** para iniciar el primer scraping.")
    st.stop()

df_lasertam = df_all[df_all["competitor"] == "Lasertam"]
df_comp     = df_all[df_all["competitor"] != "Lasertam"]

# ── Tabs ───────────────────────────────────────────────────────────────────

tab1, tab2, tab3, tab4 = st.tabs([
    "Comparación por zona",
    "Ranking competitivo",
    "Historial de precios",
    "Descuentos activos",
])

# ── TAB 1: Tabla comparativa ───────────────────────────────────────────────
with tab1:

    # Controles
    col_s, col_ses, col_ord = st.columns([3, 2, 2])
    search_q = col_s.text_input("Buscar zona…", placeholder="Ej: Axilas, Piernas…", label_visibility="collapsed")
    ses_options = sorted([s for s in df_all["sessions"].dropna().unique() if s], key=int)
    ses_sel = col_ses.selectbox(
        "Sesiones",
        options=[None] + [int(s) for s in ses_options],
        format_func=lambda x: "Todas las sesiones" if x is None else f"{x} sesión{'es' if x > 1 else ''}",
        label_visibility="collapsed",
    )
    sort_by = col_ord.selectbox(
        "Ordenar",
        ["Nombre A→Z", "Precio Lasertam ↑", "Precio Lasertam ↓"],
        label_visibility="collapsed",
    )

    # KPIs (datos Lasertam filtrados)
    df_lt_tab = df_lasertam.copy()
    if ses_sel:
        df_lt_tab = df_lt_tab[df_lt_tab["sessions"] == ses_sel]

    k1, k2, k3, k4 = st.columns(4)
    n_zones = df_all["zone_name"].nunique()
    k1.metric("Zonas mostradas", n_zones)
    if not df_lt_tab.empty:
        k2.metric("Precio mín. Lasertam", fmt_clp(df_lt_tab["price"].min()))
        k3.metric("Precio máx. Lasertam", fmt_clp(df_lt_tab["price"].max()))
        avg_disc = df_lt_tab["discount_pct"].mean()
        k4.metric("Descuento promedio",
                  f"{avg_disc:.0f}%" if pd.notna(avg_disc) else "—")
    st.divider()

    # Ordenar df
    df_tab = df_all.copy()
    if ses_sel:
        df_tab = df_tab[df_tab["sessions"] == ses_sel]

    if sort_by == "Precio Lasertam ↑":
        lt_order = df_lt_tab.groupby("zone_name")["price"].min().reset_index()
        lt_order.columns = ["zone_name", "_sort"]
        df_tab = df_tab.merge(lt_order, on="zone_name", how="left").sort_values("_sort").drop(columns="_sort")
    elif sort_by == "Precio Lasertam ↓":
        lt_order = df_lt_tab.groupby("zone_name")["price"].min().reset_index()
        lt_order.columns = ["zone_name", "_sort"]
        df_tab = df_tab.merge(lt_order, on="zone_name", how="left").sort_values("_sort", ascending=False).drop(columns="_sort")

    # Determinar sesiones a mostrar en la tabla
    if ses_sel:
        sessions_to_show = [ses_sel]
    else:
        sessions_to_show = sorted(
            [int(s) for s in df_tab["sessions"].dropna().unique()],
            key=int
        )

    coupon_info = "  |  Cupones aplicados: " + ", ".join(
        f"{c} -{pct}%" for c, pct in COUPONS.items()
    ) if COUPONS else ""
    st.caption(
        "🟢 Verde = precio más barato · 🔴 Rojo claro = precio más caro · "
        "tachado = precio sin descuento · 🏷 = precio con cupón aplicado" + coupon_info
    )

    html_table = build_comparison_table(df_tab, sessions_to_show, search=search_q)
    st.markdown(html_table, unsafe_allow_html=True)

    # ── Descarga Excel ──────────────────────────────────────────────────────
    st.markdown("<div style='margin-top:12px'></div>", unsafe_allow_html=True)
    excel_bytes = generate_excel_report(df_tab)
    from datetime import date
    st.download_button(
        label="Descargar reporte Excel",
        data=excel_bytes,
        file_name=f"precios_competencia_{date.today().isoformat()}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=False,
    )


# ── TAB 2: Ranking ─────────────────────────────────────────────────────────
with tab2:
    st.markdown("#### Posición de Lasertam frente a cada competidor")

    if df_lasertam.empty:
        st.info("No hay datos de Lasertam.")
    else:
        # ── Resumen ejecutivo por competidor ────────────────────────────────
        summary_cols = st.columns(len(df_comp["competitor"].unique()))
        for col_idx, comp_name in enumerate(sorted(df_comp["competitor"].unique())):
            df_c = df_comp[df_comp["competitor"] == comp_name]
            merged_s = df_c.merge(
                df_lasertam[["zone_name", "gender", "sessions", "price"]].rename(
                    columns={"price": "lt_price"}),
                on=["zone_name", "gender", "sessions"], how="inner",
            )
            if merged_s.empty:
                continue
            merged_s["diff_pct"] = (
                (merged_s["price"] - merged_s["lt_price"]) / merged_s["lt_price"] * 100
            )
            avg_diff   = merged_s["diff_pct"].mean()
            n_cheaper  = (merged_s["diff_pct"] > 0).sum()   # Lasertam más barato
            n_expensive = (merged_s["diff_pct"] < 0).sum()  # Lasertam más caro
            total      = len(merged_s)

            with summary_cols[col_idx]:
                if avg_diff > 0:
                    label = f"Lasertam {avg_diff:.0f}% más barato"
                    delta_color = "normal"
                elif avg_diff < 0:
                    label = f"Lasertam {abs(avg_diff):.0f}% más caro"
                    delta_color = "inverse"
                else:
                    label = "Precios iguales"
                    delta_color = "off"
                st.metric(
                    label=comp_name,
                    value=label,
                    delta=f"{n_cheaper}/{total} zonas más barato",
                    delta_color=delta_color,
                )

        st.divider()

        for comp_name in sorted(df_comp["competitor"].unique()):
            df_c = df_comp[df_comp["competitor"] == comp_name].copy()
            merged_c = df_c.merge(
                df_lasertam[["zone_name", "gender", "sessions", "price"]].rename(
                    columns={"price": "lt_price"}),
                on=["zone_name", "gender", "sessions"],
                how="inner",
            )
            if merged_c.empty:
                continue

            merged_c["diferencia"]     = merged_c["price"] - merged_c["lt_price"]
            merged_c["diferencia_pct"] = (merged_c["diferencia"] / merged_c["lt_price"] * 100).round(1)
            merged_c["status"]         = merged_c["diferencia"].apply(
                lambda d: "🟢 Lasertam más barato" if d > 0 else (
                          "🔴 Lasertam más caro"   if d < 0 else "➖ Igual"))

            n_baratas = (merged_c["diferencia"] > 0).sum()
            n_caras   = (merged_c["diferencia"] < 0).sum()

            with st.expander(
                f"Lasertam vs. **{comp_name}**  —  "
                f"🟢 {n_baratas} zonas más barato · 🔴 {n_caras} zonas más caro",
                expanded=True,
            ):
                col_a, col_b = st.columns([2, 1])
                with col_a:
                    fig = px.bar(
                        merged_c.sort_values("diferencia", ascending=False),
                        x="zone_name", y="diferencia", color="diferencia",
                        color_continuous_scale=["#dc2626", "#e5e7eb", "#16a34a"],
                        title=f"Diferencia: {comp_name} − Lasertam (CLP)",
                        labels={"diferencia": "Diferencia ($)", "zone_name": "Zona"},
                    )
                    fig.update_layout(
                        coloraxis_showscale=False,
                        xaxis_tickangle=-35,
                        height=380,
                    )
                    fig.add_hline(y=0, line_dash="dot", line_color="gray")
                    st.plotly_chart(fig, use_container_width=True)

                with col_b:
                    show = merged_c[["zone_name", "sessions", "lt_price", "price",
                                     "diferencia_pct", "status"]].copy()
                    show["lt_price"]       = show["lt_price"].apply(fmt_clp)
                    show["price"]          = show["price"].apply(fmt_clp)
                    show["diferencia_pct"] = show["diferencia_pct"].apply(lambda x: f"{x:+.1f}%")
                    show["sessions"]       = show["sessions"].apply(
                        lambda s: f"{int(s)} ses." if pd.notna(s) else "Paq.")
                    show.columns = ["Zona", "Ses.", "Lasertam", comp_name, "Dif.%", ""]
                    st.dataframe(show, use_container_width=True, hide_index=True)


# ── TAB 3: Historial ───────────────────────────────────────────────────────
with tab3:
    st.markdown("#### Evolución histórica de precios")

    all_zones = run_query(
        "SELECT DISTINCT zone_name FROM price_records ORDER BY zone_name"
    )["zone_name"].tolist()

    col_h1, col_h2, col_h3 = st.columns(3)
    zone_h    = col_h1.selectbox("Zona", all_zones, key="hist_zone")
    gender_h  = col_h2.radio("Género", ["F", "M"], key="hist_gender", horizontal=True)
    sessions_h = col_h3.selectbox(
        "Sesiones", [None, 1, 3, 6, 9], key="hist_ses",
        format_func=lambda x: "Todas" if x is None else str(x),
    )

    df_hist = load_price_history(zone_h, gender_h, sessions_h)

    if df_hist.empty:
        st.info("No hay historial aún. Los datos se acumulan diariamente cuando hay cambios.")
    else:
        view_mode = st.radio(
            "Vista", ["Precio absoluto (CLP)", "Variación % desde primer dato"],
            horizontal=True, key="hist_view",
        )

        if view_mode == "Precio absoluto (CLP)":
            fig3 = px.line(
                df_hist, x="fecha", y="price", color="competitor",
                color_discrete_map=COMPETITOR_COLORS, markers=True,
                title=f"Historial · {zone_h}",
                labels={"price": "Precio CLP", "fecha": "Fecha", "competitor": "Empresa"},
            )
            fig3.update_layout(yaxis_tickformat="$,.0f", height=400)
        else:
            # Normalizar: índice 100 = primer precio de cada empresa
            df_norm = df_hist.copy()
            base = df_norm.groupby("competitor")["price"].transform("first")
            df_norm["variacion_pct"] = ((df_norm["price"] - base) / base * 100).round(1)
            fig3 = px.line(
                df_norm, x="fecha", y="variacion_pct", color="competitor",
                color_discrete_map=COMPETITOR_COLORS, markers=True,
                title=f"Variación de precio · {zone_h}",
                labels={"variacion_pct": "Variación %", "fecha": "Fecha", "competitor": "Empresa"},
            )
            fig3.update_layout(height=400)
            fig3.add_hline(y=0, line_dash="dot", line_color="gray")
            fig3.update_traces(
                hovertemplate="%{fullData.name}<br>%{x}<br><b>%{y:+.1f}%</b><extra></extra>"
            )

        fig3.update_layout(legend=dict(orientation="h", yanchor="bottom", y=1.02))
        st.plotly_chart(fig3, use_container_width=True)

        pivot_h = df_hist.pivot_table(
            index="fecha", columns="competitor", values="price", aggfunc="min"
        ).reset_index()
        for col in pivot_h.columns[1:]:
            pivot_h[col] = pivot_h[col].apply(fmt_clp)
        st.dataframe(pivot_h, use_container_width=True, hide_index=True)


# ── TAB 4: Descuentos activos ──────────────────────────────────────────────
with tab4:
    st.markdown("#### Descuentos y promociones activas")

    df_disc = df_all[df_all["discount_pct"].notna() & (df_all["discount_pct"] > 0)].copy()

    if df_disc.empty:
        st.info("No se detectaron descuentos en el último scraping.")
    else:
        df_disc_sorted = df_disc.sort_values("discount_pct", ascending=False)

        fig4 = px.bar(
            df_disc_sorted.head(30),
            x="zone_name", y="discount_pct", color="competitor",
            color_discrete_map=COMPETITOR_COLORS, barmode="group",
            title="Top descuentos activos (%)",
            labels={"discount_pct": "Descuento %", "zone_name": "Zona", "competitor": "Empresa"},
        )
        fig4.update_layout(xaxis_tickangle=-35)
        st.plotly_chart(fig4, use_container_width=True)

        disp = df_disc_sorted[["competitor", "zone_name", "gender", "sessions",
                               "price", "original_price", "discount_pct"]].copy()
        disp["price"]          = disp["price"].apply(fmt_clp)
        disp["original_price"] = disp["original_price"].apply(fmt_clp)
        disp["discount_pct"]   = disp["discount_pct"].apply(lambda x: f"{x:.0f}%")
        disp["sessions"]       = disp["sessions"].apply(
            lambda s: f"{int(s)} ses." if pd.notna(s) else "Paquete")
        disp.columns = ["Empresa", "Zona", "Gén.", "Ses.", "Precio Oferta", "Precio Normal", "Dcto."]
        st.dataframe(disp, use_container_width=True, hide_index=True)


# ── Footer ─────────────────────────────────────────────────────────────────
st.markdown("""
<hr style="border:none;border-top:1px solid #e2e8f0;margin:32px 0 12px 0">
<p style="font-size:11px;color:#94a3b8;text-align:center;margin:0">
    PrecioSpy &nbsp;·&nbsp; Lasertam &nbsp;·&nbsp; Actualización diaria automática &nbsp;·&nbsp; Chile
</p>
""", unsafe_allow_html=True)
